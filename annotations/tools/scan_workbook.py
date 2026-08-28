#!/usr/bin/env python3
"""Scan the Market Pricing workbook for data steps and report them.

Run this after each data refresh. It is the standing check: a splice surfaces
the same week it appears rather than three years later, which is how the five
already in the file went unnoticed.

    python3 tools/scan_workbook.py "Market Pricing.xlsx"
    python3 tools/scan_workbook.py series.json          # {"gas": {"Win-26": {...}}, "power": {...}}

A step is a single day move over 15% that holds its new level for the next
five closes while the SAME delivery period in the other fuel does not move
with it. See src/stepguard.py for why that test and not a size threshold.

openpyxl is needed for the .xlsx route and is not a dependency of this
package: the daily job never opens the workbook, it is handed series. The
JSON route is stdlib only.

Exit code is 1 when anything is found, so this can gate a data refresh.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from stepguard import scan_pairs  # noqa: E402

TABS = (
    ("gas", "Gas Seasonal"),
    ("gas", "Gas Monthly"),
    ("power", "Electricity Seasonal"),
    ("power", "Electricity Monthly"),
)


def from_workbook(path: Path) -> dict[str, dict[str, dict[str, float]]]:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl is not installed, so the .xlsx route is unavailable.")
        print("Either pip install openpyxl, or pass a JSON file of series instead.")
        raise SystemExit(2)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    book: dict[str, dict[str, dict[str, float]]] = {"gas": {}, "power": {}}
    for fuel, tab in TABS:
        if tab not in wb.sheetnames:
            continue
        for period, series in _tab_series(wb[tab]).items():
            book[fuel][period] = series
    wb.close()
    return book


def _tab_series(ws) -> dict[str, dict[str, float]]:
    """Period name -> {date: close}.

    The layout is a preamble, then a row of period names, then a row of RICs
    with 'Date' in one column, then the data. The header row is found rather
    than assumed, because the preamble has changed before.
    """
    import datetime

    rows = list(ws.iter_rows(values_only=True))
    hdr = None
    for i, r in enumerate(rows[:12]):
        if r and any(isinstance(c, str) and c.strip() == "Date" for c in r):
            hdr = i
            break
    if hdr is None:
        return {}

    ric_row = rows[hdr]
    name_row = rows[hdr - 1] if hdr else [None] * len(ric_row)
    date_col = [j for j, c in enumerate(ric_row) if isinstance(c, str) and c.strip() == "Date"][0]

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    def label(cell):
        if isinstance(cell, datetime.datetime):
            return "%s-%02d" % (months[cell.month - 1], cell.year % 100)
        return str(cell).strip() if cell else ""

    cols = []
    for j in range(date_col + 1, len(ric_row)):
        ric = ric_row[j]
        if isinstance(ric, str) and ric.strip():
            cols.append((j, label(name_row[j] if j < len(name_row) else None)))

    out: dict[str, dict[str, float]] = {}
    for r in rows[hdr + 1 :]:
        if not r or len(r) <= date_col:
            continue
        d = r[date_col]
        if not isinstance(d, datetime.datetime):
            continue
        iso = d.strftime("%Y-%m-%d")
        for j, period in cols:
            if not period:
                continue
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)):
                out.setdefault(period, {})[iso] = float(v)
    return out


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__.strip().splitlines()[0])
        print("usage: python3 tools/scan_workbook.py <Market Pricing.xlsx | series.json>")
        return 2

    path = Path(argv[1])
    if not path.exists():
        print("no such file: %s" % path)
        return 2

    if path.suffix.lower() == ".json":
        raw = json.loads(path.read_text(encoding="utf-8"))
        book = {"gas": raw.get("gas", {}), "power": raw.get("power", raw.get("electricity", {}))}
    else:
        book = from_workbook(path)

    found = scan_pairs(book["gas"], book["power"])
    if not found:
        print("no uncorroborated steps found")
        return 0

    # Group by date: a splice hits a block of contracts at once, and reading
    # it contract by contract hides that.
    by_date: dict[str, list[str]] = {}
    detail: dict[str, dict] = {}
    for name, hits in found.items():
        for h in hits:
            by_date.setdefault(h["date"], []).append(name)
            detail.setdefault(h["date"] + name, h)

    print("%d date(s) with steps that the other fuel does not corroborate:" % len(by_date))
    for date in sorted(by_date):
        names = sorted(by_date[date])
        sizes = [detail[date + n]["pct"] for n in names]
        print(
            "  %s  %d contract(s), %+.1f%% to %+.1f%%"
            % (date, len(names), min(sizes), max(sizes))
        )
        print("      " + ", ".join(names[:8]) + (", ..." if len(names) > 8 else ""))
    print("")
    print("Check these against the exchange before trusting the history behind them.")
    return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
